import openpyxl
import json

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Headers:", headers)
print("Total rows:", ws.max_row - 1)
print()

data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    item = dict(zip(headers, row))
    data.append(item)

# Print unique values
brands = sorted(set(d['brand'] for d in data))
types = sorted(set(d['vehicle_type'] for d in data))
prices = [d['price_million_vnd'] for d in data]
fuels = [d['fuel_consumption_l_per_100km'] for d in data]

print("Brands:", brands)
print("Vehicle types:", types)
print(f"Price range: {min(prices):.1f} - {max(prices):.1f} million VND")
print(f"Fuel range: {min(fuels):.1f} - {max(fuels):.1f} L/100km")
print()
print("First 5 rows:")
for d in data[:5]:
    print(d)

# Also save to JSON for reference
with open('data_check.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print("\nSaved to data_check.json")
