"""
DSS Motor Choice - FastAPI Backend (v2.0)
Cải thiện: Error handling chặt chẽ, logging, validation đầy đủ, CORS, docs
"""

import logging
import os
import time
from contextlib import asynccontextmanager
from typing import List, Optional

import numpy as np
import pandas as pd
from fastapi import FastAPI, HTTPException, Request, status, Depends
from typing import List, Optional, Any, cast
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import io
from pydantic import BaseModel, Field, field_validator, model_validator
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ai_engine import (
    AHPCalculator, AHPWeightPredictor, DSSEngine,
    generate_training_data, get_dss_engine, get_predictor, initialize_system,
    _train_and_save
)

# ============================================================
# LOGGING SETUP
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger("DSS.API")


# ============================================================
# PYDANTIC SCHEMAS
# ============================================================

class UserProfile(BaseModel):
    budget: float = Field(..., ge=10, le=500, description="Ngân sách (triệu VNĐ)")
    purpose: int = Field(..., ge=0, le=3, description="0=Sinh viên, 1=Văn phòng, 2=Tour, 3=Dịch vụ")
    vehicle_type_preference: int = Field(..., ge=0, le=3, description="0=Xe số, 1=Tay ga, 2=Côn tay, 3=Tất cả")
    daily_distance_km: float = Field(..., ge=1, le=300, description="Khoảng cách đi hàng ngày (km)")
    priority_price: float = Field(..., ge=0, le=1, description="Mức độ ưu tiên giá (0–1)")
    priority_fuel: float = Field(..., ge=0, le=1, description="Mức độ ưu tiên tiết kiệm xăng (0–1)")
    priority_performance: float = Field(..., ge=0, le=1, description="Mức độ ưu tiên hiệu năng (0–1)")
    priority_design: float = Field(..., ge=0, le=1, description="Mức độ ưu tiên thiết kế (0–1)")
    priority_brand: float = Field(..., ge=0, le=1, description="Mức độ ưu tiên thương hiệu (0–1)")
    top_n: Optional[int] = Field(6, ge=1, le=20)
    min_engine_cc: Optional[int] = Field(None, ge=50, le=1000)
    max_engine_cc: Optional[int] = Field(None, ge=50, le=1000)
    brands: Optional[List[str]] = Field(None, description="Lọc theo thương hiệu cụ thể")

    @model_validator(mode='after')
    def validate_engine_cc_range(self) -> 'UserProfile':
        if self.min_engine_cc and self.max_engine_cc:
            if self.min_engine_cc > self.max_engine_cc:
                raise ValueError("min_engine_cc không được lớn hơn max_engine_cc")
        return self

    @model_validator(mode='after')
    def validate_priorities_not_all_zero(self) -> 'UserProfile':
        total = (self.priority_price + self.priority_fuel + self.priority_performance
                 + self.priority_design + self.priority_brand)
        if total <= 0:
            raise ValueError("Ít nhất một tiêu chí ưu tiên phải > 0")
        return self

    model_config = {"json_schema_extra": {
        "example": {
            "budget": 40, "purpose": 0, "vehicle_type_preference": 0,
            "daily_distance_km": 15, "priority_price": 0.35,
            "priority_fuel": 0.30, "priority_performance": 0.15,
            "priority_design": 0.10, "priority_brand": 0.10, "top_n": 6
        }
    }}


class AHPMatrixRequest(BaseModel):
    matrix: List[List[float]] = Field(..., description="Ma trận so sánh cặp NxN Saaty")
    names: Optional[List[str]] = Field(None, description="Danh sách tên các phần tử")

    @field_validator('matrix')
    @classmethod
    def validate_matrix_size(cls, v):
        n = len(v)
        if n < 2:
            raise ValueError("Ma trận phải có ít nhất 2 hàng")
        for row in v:
            if len(row) != n:
                raise ValueError(f"Mỗi hàng phải có đúng {n} phần tử")
            for val in row:
                if val <= 0:
                    raise ValueError("Tất cả giá trị trong ma trận phải > 0")
        return v


class RetrainRequest(BaseModel):
    n_samples: Optional[int] = Field(300, ge=100, le=2000)


class SensitivityRequest(BaseModel):
    """Tính lại ranking với trọng số tùy chỉnh (client-side DSS)."""
    weights: dict = Field(..., description="Dict trọng số {w_price, w_fuel, ...}")
    motorcycles: List[dict] = Field(..., description="Danh sách xe (có raw data)")

    @field_validator('weights')
    @classmethod
    def validate_weights_keys(cls, v):
        required = {'w_price', 'w_fuel', 'w_performance', 'w_design', 'w_brand'}
        missing = required - set(v.keys())
        if missing:
            raise ValueError(f"Thiếu trọng số: {missing}")
        for k, val in v.items():
            if val < 0:
                raise ValueError(f"Trọng số '{k}' không được âm")
        return v


class MotorcycleBase(BaseModel):
    brand: str
    model: str
    vehicle_type: str
    engine_cc: int
    price_million_vnd: float
    fuel_consumption_l_per_100km: float
    performance_score: float
    design_score: float
    brand_score: float
    image_url: Optional[str] = None


class MotorcycleCreate(MotorcycleBase):
    pass


class LoginRequest(BaseModel):
    username: str
    password: str

class ExportRequest(BaseModel):
    type: str # AI_RECOMMENDATION or MANUAL_AHP
    user_profile: Optional[dict] = None
    ahp_weights: Optional[dict] = None
    top_motorcycles: Optional[List[dict]] = None
    criteria_weights: Optional[List[dict]] = None
    final_ranking: Optional[List[dict]] = None
    
    # Detailed AHP data
    criteria_matrix: Optional[List[List[float]]] = None
    criteria_labels: Optional[List[str]] = None
    alternative_names: Optional[List[str]] = None
    alt_matrices: Optional[dict] = None # { criterion_id: matrix }
    
    # Detailed result data (including normalization & consistency)
    criteria_full_result: Optional[dict] = None
    alt_full_results: Optional[dict] = None # { criterion_id: { weights, details: {norm_matrix, ...} } }


# ============================================================
# LIFESPAN & APP SETUP
# ============================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifecycle management: khởi tạo khi start, cleanup khi stop."""
    logger.info("🚀 DSS Motor Choice API đang khởi động...")
    try:
        initialize_system()
        logger.info("✅ Hệ thống đã sẵn sàng!")
    except Exception as e:
        logger.critical(f"❌ CRITICAL: Không thể khởi tạo hệ thống: {e}")
        raise
    yield
    logger.info("🛑 DSS Motor Choice API đang tắt...")


app = FastAPI(
    title="Motor Choice DSS API",
    description="""
    ## 🏍️ Hệ thống Hỗ trợ Ra Quyết định Lựa chọn Xe máy

    Sử dụng **Random Forest AI** kết hợp **AHP** để đề xuất xe máy phù hợp nhất.

    ### Luồng hoạt động:
    1. Người dùng nhập thông tin cá nhân
    2. AI (Random Forest) dự đoán trọng số AHP
    3. Lọc xe theo ngân sách và loại
    4. Chuẩn hóa Min-Max theo tiêu chí Cost/Benefit
    5. Tính điểm tổng và xếp hạng
    6. Trả kết quả Top N + giải thích

    ### Tiêu chí AHP (5 tiêu chí):
    - 💰 Giá thành (Cost)
    - ⛽ Tiêu thụ nhiên liệu (Cost)
    - ⚡ Hiệu năng (Benefit)
    - 🎨 Thiết kế (Benefit)
    - 🏅 Thương hiệu (Benefit)
    """,
    version="2.0.0",
    lifespan=lifespan,
    docs_url="/docs",
    redoc_url="/redoc"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# MIDDLEWARE: Request logging & timing
# ============================================================

@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    duration = round((time.time() - start) * 1000, 1)
    logger.info(f"{request.method} {request.url.path} → {response.status_code} ({duration}ms)")
    response.headers["X-Process-Time-Ms"] = str(duration)
    return response


# ============================================================
# EXCEPTION HANDLERS
# ============================================================

@app.exception_handler(Exception)
async def generic_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception on {request.url.path}: {exc}", exc_info=True)
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"detail": "Lỗi hệ thống nội bộ. Vui lòng thử lại.", "error": str(exc)}
    )


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _get_system_or_raise():
    """Lấy predictor & dss_engine, raise 503 nếu chưa sẵn sàng."""
    predictor = get_predictor()
    dss = get_dss_engine()
    if not predictor or not dss:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Hệ thống chưa sẵn sàng. Vui lòng thử lại sau vài giây."
        )
    return predictor, dss


VEHICLE_TYPE_MAP = {0: 'Xe số', 1: 'Xe tay ga', 2: 'Xe côn tay', 3: 'Tất cả'}
PURPOSE_LABELS = ["Sinh viên", "Nhân viên văn phòng", "Người đi tour", "Người chạy dịch vụ"]


# ============================================================
# ROUTES
# ============================================================

@app.get("/", tags=["System"], summary="API Info")
async def root():
    return {
        "name": "Motor Choice DSS API",
        "version": "2.0.0",
        "docs": "/docs",
        "health": "/api/health",
        "description": "Hệ thống hỗ trợ ra quyết định lựa chọn xe máy (AI + AHP + DSS)"
    }


@app.get("/api/health", tags=["System"], summary="Kiểm tra trạng thái hệ thống")
async def health_check():
    predictor = get_predictor()
    dss = get_dss_engine()
    is_ready = predictor is not None and dss is not None and predictor.is_trained

    return {
        "status": "ok" if is_ready else "initializing",
        "model_trained": predictor.is_trained if predictor else False,
        "motorcycles_loaded": len(dss.df) if dss else 0,
        "eval_metrics": predictor.eval_metrics if predictor else {},
        "feature_importance": predictor.get_feature_importance() if predictor and predictor.is_trained else {},
        "api_version": "2.0.0"
    }


@app.get("/api/motorcycles", tags=["Data"], summary="Danh sách tất cả xe máy")
async def get_all_motorcycles(
    vehicle_type: Optional[str] = None,
    brand: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None
):
    """Lấy danh sách xe máy, hỗ trợ lọc theo type/brand/price."""
    _, dss = _get_system_or_raise()

    df = dss.df.copy()

    if vehicle_type:
        df = df[df['vehicle_type'] == vehicle_type]
    if brand:
        df = df[df['brand'].str.lower() == brand.lower()]
    if min_price is not None:
        df = df[df['price_million_vnd'] >= min_price]
    if max_price is not None:
        df = df[df['price_million_vnd'] <= max_price]

    cols = ['brand', 'model', 'vehicle_type', 'engine_cc',
            'price_million_vnd', 'fuel_consumption_l_per_100km',
            'performance_score', 'design_score', 'brand_score']
    if 'image_url' in df.columns:
        cols.append('image_url')
        
    df_safe = df[cols].replace({np.nan: None})
    data = df_safe.to_dict('records')

    return {"motorcycles": data, "total": len(data), "filtered": any([vehicle_type, brand, min_price is not None, max_price is not None])}


@app.get("/api/motorcycles/stats", tags=["Data"], summary="Thống kê dataset xe máy")
async def get_statistics():
    _, dss = _get_system_or_raise()
    return dss.get_statistics()


@app.post("/api/recommend", tags=["DSS"], summary="🎯 Tư vấn xe máy – API chính")
async def recommend_motorcycles(profile: UserProfile):
    """
    **API chính**: Nhận user profile → AI dự đoán trọng số AHP → DSS xếp hạng xe.

    Quy trình:
    1. Validate input
    2. AI predict AHP weights (Random Forest)
    3. Filter bikes by budget/type/engine_cc
    4. Min-Max normalize criteria
    5. Score = Σ wj × norm(xij)
    6. Return Top N bikes + explanation
    """
    predictor, dss = _get_system_or_raise()

    user_dict = {
        'budget': profile.budget,
        'purpose': profile.purpose,
        'vehicle_type_preference': profile.vehicle_type_preference,
        'daily_distance_km': profile.daily_distance_km,
        'priority_price': profile.priority_price,
        'priority_fuel': profile.priority_fuel,
        'priority_performance': profile.priority_performance,
        'priority_design': profile.priority_design,
        'priority_brand': profile.priority_brand,
    }

    # 2. AI predict weights
    try:
        weights = predictor.predict_weights(user_dict)
    except (ValueError, RuntimeError) as e:
        raise HTTPException(status_code=422, detail=f"Lỗi dự đoán AI: {str(e)}")

    # Sanity check on weights
    weight_sum = sum(weights.values())
    if abs(weight_sum - 1.0) > 0.01:
        logger.error(f"Predicted weights sum = {weight_sum}, expected 1.0!")
        raise HTTPException(status_code=500, detail="Lỗi nội bộ: trọng số AI không hợp lệ")

    # 3. Filter bikes
    vtype = VEHICLE_TYPE_MAP.get(profile.vehicle_type_preference, 'Tất cả')
    warning = None

    try:
        filtered_df = dss.filter_motorcycles(
            max_budget=profile.budget,
            vehicle_type=vtype if vtype != 'Tất cả' else None,
            min_engine_cc=profile.min_engine_cc,
            max_engine_cc=profile.max_engine_cc,
            brands=profile.brands
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if filtered_df.empty:
        # Fallback: bỏ filter budget, giữ loại xe
        filtered_df = dss.filter_motorcycles(
            vehicle_type=vtype if vtype != 'Tất cả' else None,
            brands=profile.brands
        )
        warning = (f"Không tìm thấy xe trong ngân sách {profile.budget}M VNĐ. "
                   f"Hiển thị {len(filtered_df)} xe phù hợp nhất với loại xe bạn chọn.")
        logger.info(f"Budget fallback triggered: {profile.budget}M → returned {len(filtered_df)} bikes")

    if filtered_df.empty:
        raise HTTPException(
            status_code=404,
            detail="Không tìm thấy xe nào phù hợp với yêu cầu. Vui lòng điều chỉnh bộ lọc."
        )

    # 4. Rank
    try:
        ranked_df = dss.rank_motorcycles(weights, filtered_df)
    except (ValueError, AssertionError) as e:
        logger.error(f"Ranking error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi xếp hạng: {str(e)}")

    top_df = ranked_df.head(profile.top_n)

    # 5. Build result
    result_records = []
    for _, row in top_df.iterrows():
        result_records.append({
            'rank': int(row['rank']),
            'brand': str(row['brand']),
            'model': str(row['model']),
            'image_url': str(row.get('image_url', '')),
            'vehicle_type': str(row['vehicle_type']),
            'engine_cc': int(row['engine_cc']),
            'price_million_vnd': round(float(row['price_million_vnd']), 2),
            'fuel_consumption_l_per_100km': round(float(row['fuel_consumption_l_per_100km']), 2),
            'performance_score': round(float(row['performance_score']), 1),
            'design_score': round(float(row['design_score']), 1),
            'brand_score': round(float(row['brand_score']), 1),
            'total_score': round(float(row['total_score']), 4),
            'scores': {
                'price_norm': round(float(row.get('norm_price_million_vnd', 0)), 4),
                'fuel_norm': round(float(row.get('norm_fuel_consumption_l_per_100km', 0)), 4),
                'performance_norm': round(float(row.get('norm_performance_score', 0)), 4),
                'design_norm': round(float(row.get('norm_design_score', 0)), 4),
                'brand_norm': round(float(row.get('norm_brand_score', 0)), 4),
            }
        })

    # 6. Explanation
    explanation = ""
    if result_records:
        try:
            explanation = dss.explain_result(top_df.iloc[0], weights)
        except Exception as e:
            logger.warning(f"Could not generate explanation: {e}")
            explanation = "Không thể tạo giải thích tự động."

    # 7. Weight contributions
    weight_labels = {
        'w_price': {'label': '💰 Giá thành', 'type': 'cost', 'description': 'Càng thấp càng tốt'},
        'w_fuel': {'label': '⛽ Tiêu thụ xăng', 'type': 'cost', 'description': 'Càng thấp càng tốt'},
        'w_performance': {'label': '⚡ Hiệu năng', 'type': 'benefit', 'description': 'Càng cao càng tốt'},
        'w_design': {'label': '🎨 Thiết kế', 'type': 'benefit', 'description': 'Càng cao càng tốt'},
        'w_brand': {'label': '🏅 Thương hiệu', 'type': 'benefit', 'description': 'Càng cao càng tốt'},
    }

    criteria_contribution = {
        info['label']: {
            'weight': round(weights[k], 4),
            'weight_pct': round(weights[k] * 100, 1),
            'type': info['type'],
            'description': info['description']
        }
        for k, info in weight_labels.items()
    }

    return {
        "success": True,
        "ahp_weights": {k: round(v, 4) for k, v in weights.items()},
        "ahp_weights_pct": {k: round(v * 100, 1) for k, v in weights.items()},
        "criteria_contribution": criteria_contribution,
        "top_motorcycles": result_records,
        "explanation": explanation,
        "total_filtered": int(len(filtered_df)),
        "total_ranked": int(len(ranked_df)),
        "warning": warning,
        "user_profile_summary": {
            "purpose_label": PURPOSE_LABELS[profile.purpose] if profile.purpose < len(PURPOSE_LABELS) else str(profile.purpose),
            "vehicle_type_label": VEHICLE_TYPE_MAP[profile.vehicle_type_preference] if profile.vehicle_type_preference < len(VEHICLE_TYPE_MAP) else str(profile.vehicle_type_preference),
            "budget": profile.budget,
            "daily_km": profile.daily_distance_km,
        }
    }


@app.post("/api/export-report", tags=["Report"], summary="Xuất báo cáo ra file Excel")
def export_report(data: ExportRequest):
    """Xuất kết quả gợi ý / AHP thành file Excel (đã được định dạng đẹp)"""
    output = io.BytesIO()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border_side = Side(border_style="thin", color="000000")
    full_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    def apply_sheet_style(worksheet, df):
        # Auto-adjust column widths and apply header styles
        for col_idx, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 4
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)
            
            # Header style
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = full_border

        # Data styles
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = full_border
                cell.alignment = Alignment(vertical="center")

    def write_ahp_block(ws, start_row, title, labels, matrix, result):
        """Builds the visual layout shifted 2 cols right with a thick black outer border."""
        c_off = 3 # Shift right by 2 cols (Start at Col 3)
        n = len(labels)
        
        # Calculate block width and height for bonding box
        # Max width is Title banner or Table 2 width (n + 5 cols)
        block_width = n + 6
        last_data_row = start_row + (n + 4) * 2 + 1 # Approximate height

        # Draw the thick black border around the whole block (roughly)
        thick_side = Side(style='medium', color="000000")
        for r in range(start_row - 1, last_data_row + 1):
            ws.cell(row=r, column=c_off - 1).border = Border(right=thick_side)
            ws.cell(row=r, column=c_off + block_width).border = Border(left=thick_side)
        for c in range(c_off - 1, c_off + block_width + 1):
            ws.cell(row=start_row - 1, column=c).border = Border(bottom=thick_side)
            ws.cell(row=last_data_row, column=c).border = Border(top=thick_side)

        # 1. Title Banner (Ultra Soft Cream/Yellow)
        # Shifted: column starts at c_off + 2
        banner_start = c_off + 1
        ws.merge_cells(start_row=start_row, start_column=banner_start, end_row=start_row, end_column=c_off + n + 4)
        cell = ws.cell(row=start_row, column=banner_start, value=title.upper())
        cell.font = Font(bold=True, size=11, color="1E3A8A")
        cell.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        cell.alignment = center_align
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        r_base = start_row + 4
        
        # Table 1: Original Matrix (Saaty)
        side_h = "TIÊU CHÍ" if "Tiêu chí" in title else "PHƯƠNG ÁN"
        h1 = ws.cell(row=r_base, column=c_off, value=side_h)
        h1.font = Font(bold=True)
        h1.fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid") # Ultra Soft Orange
        h1.border = full_border
        
        # Top Headers
        for j, lab in enumerate(labels, 1):
            c = ws.cell(row=r_base, column=c_off+j, value=lab)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
            c.border = full_border
        
        # Data for T1
        for i, lab in enumerate(labels, 1):
            row_h = ws.cell(row=r_base+i, column=c_off, value=lab)
            row_h.fill = PatternFill(start_color="F0F9FF", end_color="F0FDF4", fill_type="solid")
            row_h.border = full_border
            for j in range(1, n+1):
                val = matrix[i-1][j-1]
                if val == 1: disp = "1"
                elif val < 1: disp = f"1/{round(1/val)}"
                else: disp = f"{round(val)}"
                
                cell = ws.cell(row=r_base+i, column=c_off+j, value=disp)
                cell.border = full_border
                cell.alignment = center_align
                if i == j: 
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        # Arrow
        ws.merge_cells(start_row=r_base+n+1, start_column=c_off + n//2, end_row=r_base+n+2, end_column=c_off + n//2)
        arrow_cell = ws.cell(row=r_base+n+1, column=c_off + n//2, value="↓")
        arrow_cell.font = Font(bold=True, size=18, color="22C55E")
        arrow_cell.alignment = center_align

        # Table 2: Normalized & Result
        r_base_2 = r_base + n + 4
        h2 = ws.cell(row=r_base_2, column=c_off, value=side_h)
        h2.font = Font(bold=True)
        h2.fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
        h2.border = full_border
        
        headers_t2 = labels + ["Sum W", "Weight %", "Cons." ]
        t2_headers_colors = ["F0F9FF"]*n + ["F0FDF4", "EFF6FF", "FFFBEB"]
        
        for idx, h in enumerate(headers_t2):
            cell = ws.cell(row=r_base_2, column=c_off + 1 + idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=t2_headers_colors[idx], end_color=t2_headers_colors[idx], fill_type="solid")
            cell.border = full_border
            cell.alignment = center_align

        # Data for T2
        details = result.get('details', {})
        norm_mat = details.get('norm_matrix', [])
        sum_rows = details.get('weighted_sum_value', [])
        final_ws = result.get('weights', [])
        cons_vec = details.get('consistency_vector', [])

        for i, lab in enumerate(labels, 1):
            row_h2 = ws.cell(row=r_base_2+i, column=c_off, value=lab)
            row_h2.fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
            row_h2.border = full_border
            for j in range(n):
                c = ws.cell(row=r_base_2+i, column=c_off+1+j, value=round(norm_mat[i-1][j], 4) if norm_mat else 0)
                c.border = full_border
                c.alignment = center_align
            
            cw = ws.cell(row=r_base_2+i, column=c_off+1+n, value=round(sum_rows[i-1], 4) if sum_rows else 0)
            cw.fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            cw.border = full_border
            rw = ws.cell(row=r_base_2+i, column=c_off+2+n, value=round(final_ws[i-1], 4) if final_ws else 0)
            rw.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            rw.border = full_border
            cv = ws.cell(row=r_base_2+i, column=c_off+3+n, value=round(cons_vec[i-1], 4) if cons_vec else 0)
            cv.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
            cv.border = full_border

        ws.cell(row=r_base_2+n+1, column=c_off+1+n, value=f"λ max: {round(result.get('lambda_max',0),3)}").font = Font(bold=True)
        ws.cell(row=r_base_2+n+1, column=c_off+3+n, value=f"CR: {round(result.get('CR',0), 3)}").font = Font(bold=True)

        return r_base_2 + n + 6

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if data.type == 'AI_RECOMMENDATION':
            # 1. User Profile
            if data.user_profile:
                profile_map = {
                    "purpose_label": "Mục đích sử dụng",
                    "vehicle_type_label": "Loại xe ưu tiên",
                    "budget": "Ngân sách (Triệu VNĐ)",
                    "daily_km": "Quãng đường/ngày (Km)"
                }
                df_p = pd.DataFrame([data.user_profile]).rename(columns=profile_map)
                df_p.to_excel(writer, sheet_name='Hồ sơ người dùng', index=False)
                apply_sheet_style(writer.sheets['Hồ sơ người dùng'], df_p)

            # 2. AHP Weights
            if data.ahp_weights:
                weight_map = {
                    "w_price": "Trọng số Giá (%)",
                    "w_fuel": "Trọng số Xăng (%)",
                    "w_performance": "Trọng số Hiệu năng (%)",
                    "w_design": "Trọng số Thiết kế (%)",
                    "w_brand": "Trọng số Thương hiệu (%)"
                }
                df_w = pd.DataFrame([data.ahp_weights]).rename(columns=weight_map)
                df_w.to_excel(writer, sheet_name='Trọng số AHP', index=False)
                apply_sheet_style(writer.sheets['Trọng số AHP'], df_w)

            # 3. Top Motorcycles
            if data.top_motorcycles:
                df_bikes = pd.DataFrame(data.top_motorcycles)
                if 'image_url' in df_bikes.columns:
                    df_bikes.drop(columns=['image_url'], inplace=True)
                if 'scores' in df_bikes.columns:
                    df_bikes.drop(columns=['scores'], inplace=True)
                
                bike_map = {
                    "rank": "Xếp hạng",
                    "brand": "Thương hiệu",
                    "model": "Tên dòng xe",
                    "vehicle_type": "Phân khúc xe",
                    "engine_cc": "Dung tích (cc)",
                    "price_million_vnd": "Giá (Triệu VNĐ)",
                    "fuel_consumption_l_per_100km": "Tiêu thụ xăng (L/100km)",
                    "performance_score": "Điểm Hiệu năng/10",
                    "design_score": "Điểm Thiết kế/10",
                    "brand_score": "Điểm Thương hiệu/10",
                    "total_score": "Điểm DSS Tổng"
                }
                df_bikes = df_bikes.rename(columns=bike_map)
                df_bikes.to_excel(writer, sheet_name='Gợi ý từ AI', index=False)
                apply_sheet_style(writer.sheets['Gợi ý từ AI'], df_bikes)

        elif data.type == 'MANUAL_AHP':
            # 1. Ma trận Tiêu chí
            ws_crit = writer.book.create_sheet('Ma trận Tiêu chí', index=0)
            if data.criteria_full_result and data.criteria_matrix and data.criteria_labels:
                write_ahp_block(
                    ws_crit, 2, "Ma trận so sánh các Tiêu chí bài toán", 
                    data.criteria_labels, data.criteria_matrix, data.criteria_full_result
                )
            for i in range(1, 25): ws_crit.column_dimensions[get_column_letter(i)].width = 15

            # 2. Chi tiết tính toán PA
            ws_detail = writer.book.create_sheet('Chi tiết tính toán PA')
            if data.alt_full_results and data.alt_matrices and data.alternative_names:
                id_to_label = {
                    "price": "Giá thành", "fuel": "Tiết kiệm xăng", 
                    "performance": "Hiệu năng", "design": "Thiết kế", "brand": "Thương hiệu"
                }
                af_results = cast(dict, data.alt_full_results)
                a_matrices = cast(dict, data.alt_matrices)
                row_ptr = 2
                for crit_id, res in af_results.items():
                    matrix = a_matrices.get(crit_id)
                    if matrix:
                        label = id_to_label.get(crit_id, crit_id)
                        row_ptr = write_ahp_block(
                            ws_detail, row_ptr, f"Trọng số PA theo tiêu chí {label}",
                            cast(List[str], data.alternative_names), matrix, res
                        )
            # Finish styling detail sheet
            for i in range(1, 25): ws_detail.column_dimensions[get_column_letter(i)].width = 15

            # 3. Kết quả xếp hạng cuối cùng (Final Sheet)
            if data.final_ranking:
                df_raw = pd.DataFrame(data.final_ranking)
                if 'bike' in df_raw.columns:
                    bikes_df = pd.json_normalize(df_raw['bike'])
                    if 'image_url' in bikes_df.columns: bikes_df.drop(columns=['image_url'], inplace=True)
                    if 'scores' in bikes_df.columns: bikes_df.drop(columns=['scores'], inplace=True)
                    df_ranking = df_raw.drop(columns=['bike', 'breakdown']).join(bikes_df)
                    manual_map = {
                        "score": "Tổng điểm AHP", "rank": "Hạng", "brand": "Hãng", 
                        "model": "Model", "vehicle_type": "Loại xe", "price_million_vnd": "Giá niêm yết", 
                        "engine_cc": "Dung tích"
                    }
                    df_ranking = df_ranking.rename(columns=manual_map)
                    sheet_name = 'Kết quả xếp hạng cuối cùng'
                    df_ranking.to_excel(writer, sheet_name=sheet_name, index=False)
                    apply_sheet_style(writer.sheets[sheet_name], df_ranking)
        else:
            pd.DataFrame([{"Thông báo": "Dữ liệu không hợp lệ"}]).to_excel(writer, sheet_name='Lỗi', index=False)
            
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="DSS_Report_PhanTichXe.xlsx"'}
    )


@app.post("/api/ahp/calculate", tags=["AHP"], summary="Tính AHP từ ma trận so sánh cặp")
async def calculate_ahp(request: AHPMatrixRequest):
    """Tính trọng số AHP và Consistency Ratio từ ma trận Saaty NxN."""
    try:
        calculator = AHPCalculator()
        result = calculator.compute_weights(request.matrix, request.names)
        return result
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        logger.error(f"AHP calculation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi tính toán AHP: {str(e)}")


@app.get("/api/demo", tags=["Demo"], summary="Demo 2 người dùng mẫu")
async def run_demo():
    """Demo kết quả với Sinh viên và Nhân viên văn phòng."""
    predictor, dss = _get_system_or_raise()

    demo_profiles = [
        {
            "name": "Sinh viên – Nguyễn Văn A",
            "profile": {
                'budget': 25, 'purpose': 0, 'vehicle_type_preference': 0,
                'daily_distance_km': 15, 'priority_price': 0.35,
                'priority_fuel': 0.30, 'priority_performance': 0.15,
                'priority_design': 0.10, 'priority_brand': 0.10
            }
        },
        {
            "name": "Nhân viên văn phòng – Trần Thị B",
            "profile": {
                'budget': 70, 'purpose': 1, 'vehicle_type_preference': 1,
                'daily_distance_km': 20, 'priority_price': 0.10,
                'priority_fuel': 0.10, 'priority_performance': 0.20,
                'priority_design': 0.30, 'priority_brand': 0.30
            }
        }
    ]

    results = []
    for p in demo_profiles:
        weights = predictor.predict_weights(p['profile'])
        vtype = VEHICLE_TYPE_MAP.get(p['profile']['vehicle_type_preference'])
        filtered = dss.filter_motorcycles(
            max_budget=p['profile']['budget'],
            vehicle_type=vtype if vtype != 'Tất cả' else None
        )
        if filtered.empty:
            filtered = dss.filter_motorcycles(vehicle_type=vtype if vtype != 'Tất cả' else None)

        ranked = dss.rank_motorcycles(weights, filtered)
        top5 = ranked.head(5)

        # Xử lý NaN thành None để JSON không bị lỗi
        import numpy as np
        top5_safe = top5.replace({np.nan: None})

        results.append({
            'user': p['name'],
            'weights': {k: round(v, 3) for k, v in weights.items()},
            'top5': [
                {
                    'rank': int(row['rank']),
                    'brand': str(row['brand']),
                    'model': str(row['model']),
                    'price': round(float(row['price_million_vnd']), 1),
                    'score': round(float(row['total_score']), 4)
                }
                for _, row in top5_safe.iterrows()
            ]
        })

    return {"demo_results": results, "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")}


@app.post("/api/retrain", tags=["Admin"], summary="Huấn luyện lại model AI")
async def retrain_model(request: RetrainRequest):
    """Huấn luyện lại Random Forest với số mẫu mới (100–2000)."""
    import ai_engine

    logger.info(f"🔄 Retraining model with {request.n_samples} samples...")
    try:
        X, y = generate_training_data(request.n_samples)
        new_predictor = AHPWeightPredictor()
        metrics = new_predictor.train(X, y)

        model_path = os.path.join(os.path.dirname(__file__), '..', 'models', 'ahp_model.joblib')
        new_predictor.save(model_path)
        ai_engine._predictor = new_predictor

        logger.info(f"✅ Retrain complete. MAE={metrics['mae']}, RMSE={metrics['rmse']}")
        return {
            "success": True,
            "n_samples": request.n_samples,
            "metrics": metrics,
            "message": f"Model đã được huấn luyện lại với {request.n_samples} mẫu."
        }
    except Exception as e:
        logger.error(f"Retrain failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi huấn luyện: {str(e)}")


@app.post("/api/sensitivity/rank", tags=["DSS"], summary="Tính lại ranking với trọng số tùy chỉnh")
async def sensitivity_rank(request: SensitivityRequest):
    """
    Client-side DSS recalculation cho Sensitivity Analysis.
    Không gọi AI model – chỉ tính lại điểm với trọng số mới.
    """
    _, dss = _get_system_or_raise()

    # Normalize weights to sum = 1
    raw_weights = request.weights
    total = sum(raw_weights.values())
    if total <= 0:
        raise HTTPException(status_code=422, detail="Tổng trọng số phải > 0")

    normalized_weights = {k: v / total for k, v in raw_weights.items()}

    try:
        df = pd.DataFrame(request.motorcycles)
        ranked = dss.rank_motorcycles(normalized_weights, df)
        result = ranked[['brand', 'model', 'vehicle_type', 'price_million_vnd',
                          'total_score', 'rank']].head(10)
        return {
            "ranked": result.to_dict('records'),
            "weights_used": {k: round(v, 4) for k, v in normalized_weights.items()}
        }
    except Exception as e:
        logger.error(f"Sensitivity rank error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi tính toán: {str(e)}")


# ============================================================
# ADMIN ROUTING (CRUD)
# ============================================================

from ai_engine import DATA_PATH
from fastapi import Header

# Hardcoded for simple demo - in production use DB & Hash
ADMIN_CREDENTIALS = {"admin": "admin123"}
SECRET_TOKEN = "dss-admin-super-secret-token"

@app.post("/api/admin/login", tags=["Admin"])
async def admin_login(req: LoginRequest):
    if req.username in ADMIN_CREDENTIALS and ADMIN_CREDENTIALS[req.username] == req.password:
        return {"success": True, "token": SECRET_TOKEN}
    raise HTTPException(status_code=401, detail="Tài khoản hoặc mật khẩu không đúng")

async def verify_admin(x_admin_token: str = Header(None)):
    if x_admin_token != SECRET_TOKEN:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập trang này")
    return True

@app.get("/api/admin/motorcycles", tags=["Admin"])
async def admin_get_motorcycles(auth: bool = Depends(verify_admin)):
    _, dss = _get_system_or_raise()
    df = dss.df.copy()
    df = df.replace({np.nan: None})
    df['id'] = df.index
    return df.to_dict('records')

@app.post("/api/admin/motorcycles", tags=["Admin"])
async def admin_create_motorcycle(motor: MotorcycleCreate, auth: bool = Depends(verify_admin)):
    _, dss = _get_system_or_raise()
    new_data = motor.model_dump()
    
    # Append to dataframe
    dss.df = pd.concat([dss.df, pd.DataFrame([new_data])], ignore_index=True)
    
    # Save to Excel
    try:
        dss.df.to_excel(DATA_PATH, index=False)
        return {"success": True, "message": "Đã thêm xe mới thành công", "id": len(dss.df) - 1}
    except Exception as e:
        logger.error(f"Failed to save data: {e}")
        raise HTTPException(status_code=500, detail="Không thể lưu dữ liệu vào Excel")

@app.put("/api/admin/motorcycles/{idx}", tags=["Admin"])
async def admin_update_motorcycle(idx: int, motor: MotorcycleCreate, auth: bool = Depends(verify_admin)):
    _, dss = _get_system_or_raise()
    if idx < 0 or idx >= len(dss.df):
        raise HTTPException(status_code=404, detail="Không tìm thấy xe")
    
    update_data = motor.model_dump()
    for key, value in update_data.items():
        dss.df.at[idx, key] = value
        
    try:
        dss.df.to_excel(DATA_PATH, index=False)
        return {"success": True, "message": "Đã cập nhật thông tin xe"}
    except Exception as e:
        logger.error(f"Failed to save data: {e}")
        raise HTTPException(status_code=500, detail="Không thể lưu dữ liệu vào Excel")

@app.delete("/api/admin/motorcycles/{idx}", tags=["Admin"])
async def admin_delete_motorcycle(idx: int, auth: bool = Depends(verify_admin)):
    _, dss = _get_system_or_raise()
    if idx < 0 or idx >= len(dss.df):
        raise HTTPException(status_code=404, detail="Không tìm thấy xe")
    
    dss.df = dss.df.drop(idx).reset_index(drop=True)
    
    try:
        dss.df.to_excel(DATA_PATH, index=False)
        return {"success": True, "message": "Đã xóa xe thành công"}
    except Exception as e:
        logger.error(f"Failed to save data: {e}")
        raise HTTPException(status_code=500, detail="Không thể lưu dữ liệu vào Excel")

@app.post("/api/admin/retrain", tags=["Admin"])
async def admin_retrain(request: RetrainRequest, auth: bool = Depends(verify_admin)):
    # Re-use existing retrain logic but with auth
    return await retrain_model(request)
